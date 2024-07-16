import requests
from bs4 import BeautifulSoup
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from tempfile import NamedTemporaryFile
from openpyxl.styles import Color, PatternFill
import json
import decimal
from itertools import cycle
from twocaptcha import TwoCaptcha


def check_url(url):
    if 'https' in url:
        return url
    else:
        if 'http' in url:
            return str(url.replace('http', 'https'))
        return str('https://' + url)


def check_amazon(url):
    url = check_url(url)

    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36",
        # Add more user agents here
    ]

    user_agent_cycle = cycle(user_agents)
    try:
        headers = {
            "User-Agent": next(user_agent_cycle),
            "Accept":
            "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Referer": "https://www.google.com/",
            # Add more headers here
        }
        response = requests.get(url, headers=headers)
        # st.write(response.status_code)
        if "necesitamos asegurarnos de que no eres un robot" in response.text.lower(
        ):
            return "CAPTCHA", 0, 0, "-", "-"
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            if soup.find(string="No disponible por el momento."):
                return "INACTIVO", 0, 0, "-", "-"
            else:
                promotion_price = 0
                promotion_span = soup.find("span", class_="savingsPercentage")
                if promotion_span is not None:
                    promotion_price = soup.find("span",
                                                class_="a-price a-text-price")
                price = soup.find("span", class_="a-price-whole")
                rating = soup.find("span", "a-icon-alt")
                review = soup.find("span", id="acrCustomerReviewText")
                return ("ACTIVO", (price.text if price is not None else "-"),
                        0, (rating.text if rating is not None else "-"),
                        (review.text if review is not None else "-"))
        else:
            return "PAGINA NO ENCONTRADA", 0, 0, "-", "-"
    except requests.RequestException as e:
        st.write(e)
        return "Failed to fetch the page", 0, 0, "-", "-"


def check_mercadolibre(url):
    # url = check_url(url)
    try:
        response = requests.get(url)
        # st.write(response.status_code)
        if response.status_code == 200:
            # st.write(response.text)
            soup = BeautifulSoup(response.text, 'html.parser')
            if "publicación pausada" in response.text.lower():
                return "INACTIVO", 0, 0, "-", "-"
            else:
                price = soup.find("span",
                                  class_="andes-money-amount__fraction")
                promotion_price = soup.find(
                    "span", class_="andes-money-amount__fraction")
                price_arr = soup.find_all(
                    "span", class_="andes-money-amount__fraction")
                try:
                    list_price = price_arr[0]
                    promotion_price = None
                    if len(price_arr) > 1:
                        promotion_price = price_arr[1]
                except IndexError:
                    return "Información de precio no encontrado", 0, 0, "-", "-"

                rating = soup.find("span", "ui-pdp-review__rating")
                review = soup.find(
                    "p",
                    class_="ui-review-ui-review-capability__rating__label")
                return ("ACTIVO",
                        (list_price.text if price is not None else "-"),
                        (promotion_price.text
                         if promotion_price is not None else "-"),
                        (rating.text if rating is not None else "-"),
                        (review.text if review is not None else "-"))
        else:
            return "PAGINA NO ENCONTRADA", 0, 0, "-", "-"
    except requests.RequestException as e:
        print(e)
        return "Error al intentar acceder a la pag", 0, 0, "-", "-"


def check_walmart(url):
    try:
        response = requests.get(url)
        st.write(response.status_code)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            # Check specific indicators of availability
            if "agotado" in response.text.lower():
                st.write("INACTIVO")
                return "INACTIVO", 0, 0, "-", "-"
            else:
                st.write("Activo")
                list_price = soup.find('span', itemprop='name')
                promotion_price = None
                rating = None
                review = None
                return ("ACTIVO",
                        (list_price.text if list_price is not None else "-"),
                        (promotion_price.text
                         if promotion_price is not None else "-"),
                        (rating.text if rating is not None else "-"),
                        (review.text if review is not None else "-"))
    except requests.RequestException:
        return "Failed to fetch the page"


def check_liverpool(url):
    url = check_url(url)
    # st.write(url)
    try:
        headers = {
            "User-Agent":
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept":
            "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate, br",
            "Upgrade-Insecure-Requests": "1",
            "Referer": "https://www.google.com/",
            # Add more headers here
        }
        response = requests.get(url, headers=headers)
        # st.write(response.status_code)
        if response.status_code == 200:
            # st.write(response.text)
            soup = BeautifulSoup(response.text, 'html.parser')
            script_tag = soup.find("script", id="__NEXT_DATA__")
            json_object = json.loads(script_tag.text)
            discount_price = 0
            regular_price = 0
            if json_object["query"]["data"]["mainContent"]["records"][0][
                    "allMeta"]["variants"][0]["prices"][
                        "promoPrice"] is not None or decimal.Decimal(
                            json_object["query"]["data"]["mainContent"]
                            ["records"][0]["allMeta"]["variants"][0]["prices"]
                            ["promoPrice"]) > 0:
                discount_price = json_object["query"]["data"]["mainContent"][
                    "records"][0]["allMeta"]["variants"][0]["prices"][
                        "promoPrice"]

            regular_price = json_object["query"]["data"]["mainContent"][
                "records"][0]["allMeta"]["variants"][0]["prices"]["listPrice"]

            return ("ACTIVO",
                    (regular_price if regular_price is not None else "-"),
                    (discount_price if discount_price is not None else "-"),
                    "-", "-")
        else:
            return "INACTIVO", 0, 0, "-", "-"

    except requests.RequestException as e:
        print(e)
        return "PAGINA NO ENCONTRADA", 0, 0, "-", "-"


# def check_coppel(url):
#     try:
#         response = requests.get(url)
#         if response.status_code == 404:
#             return "Product link not available"
#         else:
#             return "ACTIVO"
#     except requests.RequestException:
#         return "Failed to fetch the page"


def check_home_depot(url):
    try:

        last_part = url.split("/")[-1]
        last_value = last_part.split("-")[-1]
        home_depot_request = f'https://www.homedepot.com.mx/search/resources/api/v2/products?langId=-5&storeId=10351&contractId=4000000000000000003&langId=-5&partNumber={last_value}&physicalStoreId=8702'
        response = requests.get(home_depot_request)
        data = response.json()
        promotion_price = None
        rating = None
        review = None
        list_price = None
        status = "INACTIVO"
        base_data = data["contents"][0]
        for price_pos in base_data["price"]:
            if price_pos["usage"] == "Offer":
                promotion_price = price_pos["value"]
            if price_pos["usage"] == "Display":
                list_price = price_pos["value"]
        if "x_ratings.total_reviews" in base_data:
            review = base_data["x_ratings.total_reviews"]
        if "x_ratings.rating" in base_data:
            rating = base_data["x_ratings.rating"]
        product_id = base_data["id"]
        status_request = f'https://www.homedepot.com.mx/wcs/resources/store/10351/inventoryavailability/{product_id}?&langId=-5&onlineStoreId=10351&search=2&physicalStoreId=12605'
        response = requests.get(status_request)
        availabilty_data = response.json()
        if availabilty_data["InventoryAvailability"][0][
                "inventoryStatus"] == "Available":
            status = "ACTIVO"

        return (status, (list_price if list_price is not None else "-"),
                (promotion_price if promotion_price is not None else "-"),
                (rating if rating is not None else "-"),
                (review if review is not None else "-"))
    except requests.RequestException:
        return "Failed to fetch the page"


def main():
    # mas de un archivo
    # descarga del zip creado de facturapi
    st.title("Marketplace Product Status Extractor")

    dataset = st.file_uploader("Upload Excel file (.xlsx)", type=['xlsx'])
    results = {}
    if dataset is not None:
        wb = openpyxl.load_workbook(dataset, read_only=True)
        st.info(f"File uploaded: {dataset.name}")

        ws = wb.active

        ###
        # End Result Excel Variables
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active

        keys = [
            "Marketplace", "Codigo", "Descripcion", "Link", "Estatus",
            "Precio Regular", "Precio Promoción", "Calificación", "# Reseñas"
        ]
        result_row_num = 1

        for col_num, column_title in enumerate(keys, 1):
            cell = result_ws.cell(row=result_row_num, column=col_num)
            cell.value = column_title

        ###
        st.write("Procesando archivo...")
        i = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            result_row_num += 1

            marketplace = row[0]
            product_code = row[1]
            product_name = row[2]
            link = row[3]

            result = ""
            price = "-"
            rating = "-"
            reviews = "-"
            promotion_price = "-"
            if marketplace == 'Amazon':
                result, price, promotion_price, rating, reviews = check_amazon(
                    link)

            elif marketplace == 'ML':
                result, price, promotion_price, rating, reviews = check_mercadolibre(
                    link)

            elif marketplace == 'Liverpool':
                result, price, promotion_price, rating, reviews = check_liverpool(
                    link)
            elif marketplace == 'Walmart':
                result, price, promotion_price, rating, reviews = check_walmart(
                    link)
            elif marketplace == 'HomeDepot':
                result, price, promotion_price, rating, reviews = check_home_depot(
                    link)

            row = [
                marketplace, product_code, product_name, link, result, price,
                promotion_price, rating, reviews
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = result_ws.cell(row=result_row_num, column=col_num)
                cell.value = cell_value
            i += 1
        st.write("Terminando de procesar archivo...")
        st.write("Se procesaron ", i, " productos.")
        for e_column in result_ws['E']:
            if e_column.value == "ACTIVO":
                e_column.fill = PatternFill(start_color='38B856',
                                            end_color='38B856',
                                            fill_type='solid')
            if e_column.value == "INACTIVO":
                e_column.fill = PatternFill(start_color='d30000',
                                            end_color='d30000',
                                            fill_type='solid')
            if e_column.value in [
                    "PAGINA NO ENCONTRADA", "Failed to fetch the page"
            ]:
                e_column.fill = PatternFill(start_color='808080',
                                            end_color='808080',
                                            fill_type='solid')
        with NamedTemporaryFile() as tmp:

            result_wb.save(tmp.name)
            data = BytesIO(tmp.read())

        st.subheader("Resultados")
        st.download_button("Descargar Archivo",
                           data=data,
                           mime='xlsx',
                           file_name="resultados.xlsx")


if __name__ == "__main__":
    main()
